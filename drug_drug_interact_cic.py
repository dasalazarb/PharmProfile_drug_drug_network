"""Herramientas para detectar interacciones entre fármacos.

Este módulo procesa un archivo de Excel con registros de administración
farmacológica por paciente y fecha, genera una representación XML y
calcula combinaciones de medicamentos en distintas ventanas de tiempo.
También ofrece una búsqueda básica de interacciones en un archivo
DrugBank.
"""
from __future__ import annotations

import argparse
import logging
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable, List, Sequence, Tuple
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.etree.ElementTree import Comment, Element, SubElement

# Rutas por defecto
DEFAULT_EXCEL_PATH = Path("Med_vs_Tiempo.xlsx")
DEFAULT_SHEET_NAME = "Med_vs_Tiempo (5)"
DEFAULT_OUTPUT_DIR = Path(".")
DEFAULT_DRUGBANK_XML = Path("drugbank_2.xml")
DEFAULT_PROFILE_LIST = Path("lista_med_cic.txt")

LOGGER = logging.getLogger(__name__)


@dataclass
class MedicationEvent:
    """Representa la administración de un medicamento en una fecha concreta."""

    timestamp: datetime
    patient: str
    medications: List[str]


def combine_date_time(date_value, time_value) -> datetime:
    """Convierte valores de fecha y hora en un objeto ``datetime``.

    Se aceptan instancias de ``datetime`` y valores de texto con formatos
    reconocidos por :func:`datetime.fromisoformat`.
    """

    if isinstance(date_value, datetime):
        date_part = date_value.date()
    else:
        date_part = datetime.fromisoformat(str(date_value)).date()

    if isinstance(time_value, datetime):
        time_part = time_value.time()
    else:
        # openpyxl suele devolver valores ``datetime`` para celdas de hora
        time_part = datetime.fromisoformat(str(time_value)).time()

    return datetime.combine(date_part, time_part)


def load_schedule(excel_path: Path, sheet_name: str) -> List[MedicationEvent]:
    """Carga el plan de medicación desde un archivo de Excel."""

    LOGGER.info("Leyendo hoja '%s' de %s", sheet_name, excel_path)
    wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    events: List[MedicationEvent] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None or row[1] is None or row[2] is None:
            continue
        date_value, time_value, patient, medication_cell = row[:4]
        medications = _split_medications(str(medication_cell)) if medication_cell else []
        try:
            timestamp = combine_date_time(date_value, time_value)
        except Exception as exc:  # pragma: no cover - robustez frente a datos sucios
            LOGGER.warning("No se pudo interpretar la fecha/hora %s %s: %s", date_value, time_value, exc)
            continue
        events.append(MedicationEvent(timestamp=timestamp, patient=str(patient), medications=medications))

    LOGGER.info("Cargados %s eventos de medicación", len(events))
    return events


def _split_medications(value: str) -> List[str]:
    """Separa combinaciones de medicamentos indicadas con guiones bajos."""

    return [med.strip() for med in value.split("_") if med.strip()]


def build_drug_tree(events: Sequence[MedicationEvent]) -> Element:
    """Construye el árbol XML con pacientes, fechas y medicamentos."""

    root = Element("drug")
    root.append(Comment("Relacion med_med por paciente por fecha"))

    patients: dict[str, Element] = {}
    for event in events:
        patient = patients.get(event.patient)
        if patient is None:
            patient = SubElement(root, event.patient, name=event.patient)
            patients[event.patient] = patient

        date_key = event.timestamp.strftime("%Y-%m-%d-%H-%M-%S")
        existing_date = patient.find(date_key)
        if existing_date is None:
            existing_date = SubElement(
                patient,
                date_key,
                name=event.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            )

        for med in event.medications:
            SubElement(existing_date, med, name=med)

    return root


def _iter_patient_dates(patient: Element) -> Iterable[Tuple[datetime, Element]]:
    for date_node in list(patient):
        try:
            yield datetime.strptime(date_node.attrib["name"], "%Y-%m-%d %H:%M:%S"), date_node
        except (KeyError, ValueError):
            LOGGER.debug("Fecha inválida en nodo %s", date_node.tag)


def _sorted_pair(name_a: str, name_b: str) -> str:
    return "_".join(sorted([name_a, name_b], key=str.lower))


@dataclass
class CombinationResults:
    within_24h: List[str]
    within_48h: List[str]
    within_6h: List[str]


def compute_time_window_combinations(drug_tree: Element) -> CombinationResults:
    """Calcula combinaciones de medicamentos en ventanas de 6h, 24h y 48h."""

    combos_24: List[str] = []
    combos_48: List[str] = []
    combos_6: List[str] = []

    for patient in list(drug_tree)[1:]:  # se omite el comentario inicial
        date_nodes = list(_iter_patient_dates(patient))
        for date_a, node_a in date_nodes:
            for date_b, node_b in date_nodes:
                _append_combinations(date_a, node_a, date_b, node_b, combos_24, combos_48, combos_6)

    return CombinationResults(within_24h=combos_24, within_48h=combos_48, within_6h=combos_6)


def _append_combinations(
    date_a: datetime,
    node_a: Element,
    date_b: datetime,
    node_b: Element,
    combos_24: List[str],
    combos_48: List[str],
    combos_6: List[str],
) -> None:
    plus_six = date_a + timedelta(hours=6)
    minus_six = date_a - timedelta(hours=6)

    meds_a = [child.attrib.get("name", child.tag) for child in list(node_a)]
    meds_b = [child.attrib.get("name", child.tag) for child in list(node_b)]

    if date_a.date() == date_b.date():
        for med_a in meds_a:
            for med_b in meds_b:
                if med_a != med_b:
                    combos_24.append(_sorted_pair(med_a, med_b))
    elif date_a + timedelta(days=2) == date_b:
        for med_a in meds_a:
            for med_b in meds_b:
                if med_a != med_b:
                    combos_48.append(_sorted_pair(med_a, med_b))
    elif minus_six < date_b < plus_six:
        for med_a in meds_a:
            for med_b in meds_b:
                if med_a != med_b:
                    combos_6.append(_sorted_pair(med_a, med_b))


def write_combinations(results: CombinationResults, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    _write_list(output_dir / "combinaciones_24h_noSorted.txt", results.within_24h)
    _write_list(output_dir / "combinaciones_48_noSorted.txt", results.within_48h)
    _write_list(output_dir / "combinaciones_6_noSorted.txt", results.within_6h)


def _write_list(path: Path, values: Sequence[str]) -> None:
    LOGGER.info("Escribiendo %s entradas en %s", len(values), path)
    path.write_text("\n".join(values))


def collect_same_day_pairs(drug_tree: Element) -> List[str]:
    """Devuelve combinaciones de medicamentos administrados el mismo día y paciente."""

    pairs: List[str] = []
    for patient in list(drug_tree)[1:]:
        for _, date_node in _iter_patient_dates(patient):
            meds = [child.attrib.get("name", child.tag) for child in list(date_node)]
            for med_a in meds:
                for med_b in meds:
                    if med_a != med_b:
                        pairs.append(_sorted_pair(med_a, med_b))
    return pairs


def load_profile_list(profile_path: Path) -> List[str]:
    if not profile_path.exists():
        LOGGER.warning("No se encontró el archivo de perfil %s", profile_path)
        return []
    return [line.strip() for line in profile_path.read_text().splitlines() if line.strip()]


def find_drugbank_interactions(profile: Sequence[str], drugbank_xml: Path) -> List[str]:
    if not profile:
        return []
    if not drugbank_xml.exists():
        LOGGER.warning("No se encontró el archivo DrugBank %s", drugbank_xml)
        return []

    root = ET.parse(drugbank_xml).getroot()
    ns = {"db": "http://www.drugbank.ca"}
    interactions: List[str] = []

    for drugbank_entry in root.findall("db:drug", ns):
        drug_name = (drugbank_entry.findtext("db:name", default="", namespaces=ns) or "").lower()
        synonyms = [syn.text.lower() for syn in drugbank_entry.findall("db:synonyms/db:synonym", ns) if syn.text]
        interaction_nodes = drugbank_entry.findall("db:drug-interactions/db:drug-interaction", ns)

        for medicine in profile:
            med_lower = medicine.lower()
            if med_lower not in drug_name and all(med_lower not in synonym for synonym in synonyms):
                continue

            for node in interaction_nodes:
                counterpart = node.findtext("db:name", default="", namespaces=ns)
                description = node.findtext("db:description", default="", namespaces=ns)
                if not counterpart:
                    continue
                pair = _sorted_pair(counterpart.lower(), drug_name)
                interactions.append(f"{pair}\t{description.lower()}")

    return interactions


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Detectar interacciones medicamento-medicamento")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH, help="Archivo Excel con el plan de medicación")
    parser.add_argument("--sheet", type=str, default=DEFAULT_SHEET_NAME, help="Nombre de la hoja en el Excel")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directorio para guardar resultados")
    parser.add_argument("--drugbank", type=Path, default=DEFAULT_DRUGBANK_XML, help="Archivo XML de DrugBank")
    parser.add_argument("--profile", type=Path, default=DEFAULT_PROFILE_LIST, help="Lista de medicamentos de interés")
    parser.add_argument(
        "--log-level",
        type=str,
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Nivel de detalle del registro",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()
    logging.basicConfig(level=getattr(logging, args.log_level))

    events = load_schedule(args.excel, args.sheet)
    if not events:
        LOGGER.error("No se encontraron eventos de medicación. Abortando.")
        return

    drug_tree = build_drug_tree(events)

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    xml_path = output_dir / "drug.xml"
    ET.ElementTree(drug_tree).write(xml_path, encoding="utf-8", xml_declaration=True)
    LOGGER.info("Archivo XML guardado en %s", xml_path)

    combinations = compute_time_window_combinations(drug_tree)
    write_combinations(combinations, output_dir)

    same_day_pairs = collect_same_day_pairs(drug_tree)
    _write_list(output_dir / "intreacciones_cic_no_depurado.txt", same_day_pairs)

    profile = load_profile_list(args.profile)
    drugbank_interactions = find_drugbank_interactions(profile, args.drugbank)
    if drugbank_interactions:
        _write_list(output_dir / "interacciones_drugbank.txt", drugbank_interactions)
    else:
        LOGGER.info("No se generó archivo de interacciones DrugBank (sin perfil o archivo faltante)")


if __name__ == "__main__":
    main()
