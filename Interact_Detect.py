"""
Revised Interact_Detect pipeline for generating drug administration combinations
and detecting DrugBank interactions.

The script processes an Excel file with patient medication administrations,
constructs an XML tree, emits combination lists for several time windows, and
cross-references a DrugBank XML export to enumerate known interactions for a
profile of interest.
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Set, Tuple

import xml.etree.ElementTree as ET
from openpyxl import load_workbook


# Global defaults
DEFAULT_EXCEL = Path("Med_vs_Tiempo.xlsx")
DEFAULT_SHEET = "Med_vs_Tiempo (5)"
DEFAULT_DRUGBANK = Path("drugbank_2.xml")
DEFAULT_PROFILE = Path("lista_med_cic.txt")
DEFAULT_OUTPUT_DIR = Path(".")
XML_FILENAME = "drug.xml"
OUTPUT_24H = "combinaciones_24h.txt"
OUTPUT_48H = "combinaciones_48h.txt"
OUTPUT_6H = "combinaciones_6h.txt"
OUTPUT_INTERACTIONS = "interacciones_drugbank.txt"


logger = logging.getLogger(__name__)


@dataclass
class Administration:
    timestamp: datetime
    medications: List[str]


def load_profile(profile_path: Path) -> List[str]:
    if not profile_path.exists():
        raise FileNotFoundError(f"Profile file not found: {profile_path}")
    with profile_path.open("r", encoding="utf-8") as profile_file:
        medications = [line.strip() for line in profile_file if line.strip()]
    logger.debug("Loaded %d medications from profile", len(medications))
    return medications


def load_administrations(excel_path: Path, sheet_name: str) -> Dict[str, List[Administration]]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    workbook = load_workbook(filename=excel_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {excel_path}")

    sheet = workbook[sheet_name]
    patients: Dict[str, List[Administration]] = defaultdict(list)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        date_cell, time_cell, patient, medication_cell = row[:4]
        if not (date_cell and time_cell and patient and medication_cell):
            logger.debug("Skipping incomplete row: %s", row)
            continue

        if isinstance(date_cell, datetime):
            date_value = date_cell.date()
        elif isinstance(date_cell, date):
            date_value = date_cell
        else:
            logger.debug("Unrecognized date value %s", date_cell)
            continue

        if isinstance(time_cell, datetime):
            time_value = time_cell.time()
        elif isinstance(time_cell, time):
            time_value = time_cell
        else:
            try:
                time_value = datetime.strptime(str(time_cell).replace("1900-01-01 ", ""), "%H:%M:%S").time()
            except ValueError:
                logger.debug("Unrecognized time value %s", time_cell)
                continue

        timestamp = datetime.combine(date_value, time_value)
        medications = [m.strip() for m in str(medication_cell).split("_") if m.strip()]
        patients[str(patient)].append(Administration(timestamp=timestamp, medications=medications))

    logger.info("Loaded administrations for %d patients", len(patients))
    return patients


def build_xml_tree(patients: Dict[str, List[Administration]]) -> ET.ElementTree:
    root = ET.Element("drug")
    root.append(ET.Comment("Relacion med_med por paciente por fecha"))

    for patient_id, administrations in patients.items():
        patient_element = ET.SubElement(root, patient_id, name=patient_id)
        for admin in sorted(administrations, key=lambda adm: adm.timestamp):
            date_key = admin.timestamp.strftime("%Y-%m-%d-%H-%M-%S")
            date_value = admin.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            date_element = ET.SubElement(patient_element, date_key, name=date_value)
            for med in admin.medications:
                ET.SubElement(date_element, med, name=med)

    return ET.ElementTree(root)


def _pairwise_medications(first: Sequence[str], second: Sequence[str]) -> Iterable[Tuple[str, str]]:
    for med_a in first:
        for med_b in second:
            if med_a.lower() == med_b.lower():
                continue
            yield tuple(sorted((med_a, med_b), key=str.lower))


def compute_combinations(patients: Dict[str, List[Administration]]) -> Tuple[Set[str], Set[str], Set[str]]:
    combos_24: Set[str] = set()
    combos_48: Set[str] = set()
    combos_6: Set[str] = set()

    for administrations in patients.values():
        sorted_admins = sorted(administrations, key=lambda adm: adm.timestamp)
        for i, admin_a in enumerate(sorted_admins):
            for admin_b in sorted_admins[i:]:
                if admin_a is admin_b:
                    continue
                date_a = admin_a.timestamp.date()
                date_b = admin_b.timestamp.date()
                delta_days = (date_b - date_a).days
                delta_hours = admin_b.timestamp - admin_a.timestamp

                if delta_days == 0:
                    target_set = combos_24
                elif delta_days == 2:
                    target_set = combos_48
                elif timedelta(hours=-6) < delta_hours < timedelta(hours=6):
                    target_set = combos_6
                else:
                    continue

                for med_pair in _pairwise_medications(admin_a.medications, admin_b.medications):
                    target_set.add("_".join(med_pair))

    combos_48.update(combos_24)
    return combos_24, combos_48, combos_6


def write_list(output_path: Path, values: Iterable[str]) -> None:
    output_path.write_text("\n".join(sorted(values)), encoding="utf-8")
    logger.info("Wrote %s", output_path)


def _drugbank_namespace(root: ET.Element) -> str:
    if root.tag.startswith("{"):
        return root.tag.split("}")[0].strip("{")
    return ""


def find_interactions(drugbank_path: Path, profile: Sequence[str]) -> Tuple[Set[str], Set[str]]:
    if not drugbank_path.exists():
        raise FileNotFoundError(f"DrugBank XML not found: {drugbank_path}")

    tree = ET.parse(drugbank_path)
    root = tree.getroot()
    namespace = _drugbank_namespace(root)
    ns = {"db": namespace} if namespace else {}

    interactions: Set[str] = set()
    not_found: Set[str] = set()

    for drug in profile:
        query = drug.lower()
        matched = False
        for entry in root:
            name_elem = entry.find("db:name", ns) if ns else entry.find("name")
            if name_elem is None or not name_elem.text:
                continue
            primary_name = name_elem.text.lower()
            synonym_elems = entry.findall("db:synonyms/db:synonym", ns) if ns else entry.findall("synonyms/synonym")
            synonyms = [syn.text.lower() for syn in synonym_elems if syn is not None and syn.text]

            if query == primary_name or query in synonyms:
                matched = True
                interaction_section = entry.find("db:drug-interactions", ns) if ns else entry.find("drug-interactions")
                if interaction_section is None:
                    continue
                for interaction in interaction_section:
                    partner = interaction.find("db:name", ns) if ns else interaction.find("name")
                    description = interaction.find("db:description", ns) if ns else interaction.find("description")
                    if partner is None or not partner.text or description is None or not description.text:
                        continue
                    ordered_pair = "_".join(sorted((query, partner.text.lower()), key=str.lower))
                    interactions.add(f"{ordered_pair}\t{description.text.strip()}")
                break
        if not matched:
            not_found.add(query)

    return interactions, not_found


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format="%(asctime)s [%(levelname)s] %(message)s")


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate drug combinations and DrugBank interactions")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to Med_vs_Tiempo Excel file")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Worksheet name inside the Excel file")
    parser.add_argument("--drugbank", type=Path, default=DEFAULT_DRUGBANK, help="Path to DrugBank XML export")
    parser.add_argument("--profile", type=Path, default=DEFAULT_PROFILE, help="Text file with one medication per line")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory to write outputs")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    configure_logging(args.verbose)

    try:
        profile = load_profile(args.profile)
        patients = load_administrations(args.excel, args.sheet)
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to load inputs: %s", exc)
        return 1

    xml_tree = build_xml_tree(patients)
    xml_output = args.output_dir / XML_FILENAME
    xml_tree.write(xml_output, encoding="utf-8", xml_declaration=True)
    logger.info("Wrote %s", xml_output)

    combos_24, combos_48, combos_6 = compute_combinations(patients)
    write_list(args.output_dir / OUTPUT_24H, combos_24)
    write_list(args.output_dir / OUTPUT_48H, combos_48)
    write_list(args.output_dir / OUTPUT_6H, combos_6)

    try:
        interactions, not_found = find_interactions(args.drugbank, profile)
        write_list(args.output_dir / OUTPUT_INTERACTIONS, interactions)
        if not_found:
            logger.warning("Medications not found in DrugBank: %s", ", ".join(sorted(not_found)))
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Failed to evaluate DrugBank interactions: %s", exc)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
